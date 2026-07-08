"""Tests for the server-side Excel export (`db.excel.build_portfolio_workbook`)."""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from db.models import AssetAllocation, Product


def _product(**overrides) -> Product:
    defaults = dict(
        user_id="usr_test",
        name="Fund A",
        provider="Provider A",
        amount=10_000,
        category="directas",
        composition=[AssetAllocation(name="Fund A", percentage=100)],
    )
    defaults.update(overrides)
    return Product(**defaults)


@pytest.fixture
def sample_products() -> list[Product]:
    return [
        _product(name="Accionariado XYZ", amount=20_000, category="directas"),
        _product(name="Deuda Privada Fund", amount=15_000, category="privados"),
        _product(
            name="Multi-asset Fund",
            amount=5_000,
            category="publicos",
            composition=[
                AssetAllocation(name="RV US Large Cap", percentage=60),
                AssetAllocation(name="RF Corporate", percentage=40),
            ],
        ),
    ]


def test_build_portfolio_workbook_returns_readable_buffer(sample_products):
    from db.excel import build_portfolio_workbook

    buffer = build_portfolio_workbook(sample_products)
    wb = load_workbook(buffer)

    assert wb is not None


def test_summary_sheet_is_first_and_named_portafolio_final(sample_products):
    from db.excel import build_portfolio_workbook

    buffer = build_portfolio_workbook(sample_products)
    wb = load_workbook(buffer)

    assert wb.sheetnames[0] == "Portafolio Final"


def test_sheet_names_include_one_per_used_category(sample_products):
    from db.excel import build_portfolio_workbook

    buffer = build_portfolio_workbook(sample_products)
    wb = load_workbook(buffer)

    assert "Inversiones directas" in wb.sheetnames
    assert "Mercados privados" in wb.sheetnames
    assert "Mercados públicos" in wb.sheetnames
    # No products in these categories — no sheet should be created
    assert "Club deals" not in wb.sheetnames
    assert "Otros" not in wb.sheetnames
    assert "Cash y equivalentes" not in wb.sheetnames


def test_category_sheet_row_count_matches_products_plus_total(sample_products):
    from db.excel import build_portfolio_workbook

    buffer = build_portfolio_workbook(sample_products)
    wb = load_workbook(buffer)

    ws = wb["Inversiones directas"]
    # header row + 1 product row + total row = 3
    assert ws.max_row == 3
    assert ws.cell(row=1, column=1).value == "Nombre"
    assert ws.cell(row=2, column=1).value == "Accionariado XYZ"
    assert ws.cell(row=2, column=3).value == 20_000
    assert ws.cell(row=3, column=1).value == "Total"
    assert ws.cell(row=3, column=3).value == 20_000


def test_summary_sheet_totals_match_input_products(sample_products):
    from db.excel import build_portfolio_workbook

    buffer = build_portfolio_workbook(sample_products)
    wb = load_workbook(buffer)

    ws = wb["Portafolio Final"]
    total = sum(p.amount for p in sample_products)

    # Last row is the grand total row
    last_row = ws.max_row
    assert ws.cell(row=last_row, column=1).value == "Total"
    assert ws.cell(row=last_row, column=2).value == total
    assert ws.cell(row=last_row, column=4).value == len(sample_products)


def test_empty_portfolio_produces_only_summary_sheet():
    from db.excel import build_portfolio_workbook

    buffer = build_portfolio_workbook([])
    wb = load_workbook(buffer)

    assert wb.sheetnames == ["Portafolio Final"]
    ws = wb["Portafolio Final"]
    assert ws.cell(row=2, column=1).value == "Total"
    assert ws.cell(row=2, column=2).value == 0


def test_export_filename_has_xlsx_extension_and_date():
    from datetime import date

    from db.excel import export_filename

    filename = export_filename()

    assert filename.startswith("portafolio-sabbi-")
    assert filename.endswith(".xlsx")
    assert date.today().isoformat() in filename
