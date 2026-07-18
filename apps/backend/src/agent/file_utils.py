"""Helpers for converting uploaded file attachments into content blocks
that the Anthropic API accepts.  Claude's ``document`` type only supports
``application/pdf``; other file types must be parsed server-side and sent
as text.
"""

from __future__ import annotations

import base64
import io
import zipfile
from xml.etree import ElementTree

SPREADSHEET_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/csv",
}

WORD_MIMES = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/msword",
}

PARSEABLE_MIMES = SPREADSHEET_MIMES | WORD_MIMES


def file_to_text(data_b64: str, mime: str) -> str | None:
    """Parse a base64-encoded file and return a text representation.
    Returns None for unsupported/unparseable formats."""
    if mime in SPREADSHEET_MIMES:
        return _spreadsheet_to_text(data_b64, mime)
    if mime in WORD_MIMES:
        return _word_to_text(data_b64, mime)
    return None


def spreadsheet_to_text(data_b64: str, mime: str) -> str | None:
    """Public alias kept for backward compatibility."""
    return _spreadsheet_to_text(data_b64, mime)


def _spreadsheet_to_text(data_b64: str, mime: str) -> str | None:
    if mime not in SPREADSHEET_MIMES:
        return None

    try:
        raw = base64.b64decode(data_b64)
    except Exception:
        return "[Error: no se pudo decodificar el archivo]"

    if mime == "text/csv":
        text = raw.decode("utf-8", errors="replace")
        return f"[Contenido del archivo CSV]\n\n{text}"

    try:
        from openpyxl import load_workbook
    except ImportError:
        return "[Error: openpyxl not installed — cannot parse Excel file]"

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"--- Hoja: {sheet.title} ---\n" + "\n".join(rows))
    wb.close()
    return (
        "[Contenido del archivo Excel]\n\n" + "\n\n".join(parts)
        if parts
        else "[Archivo Excel vacío]"
    )


_WORDML_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


def _word_to_text(data_b64: str, mime: str) -> str | None:
    if mime not in WORD_MIMES:
        return None

    if mime == "application/msword":
        return "[Archivo .doc detectado — solo se soporta .docx]"

    try:
        raw = base64.b64decode(data_b64)
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            xml_bytes = zf.read("word/document.xml")
    except Exception:
        return "[Error: no se pudo leer el archivo Word]"

    tree = ElementTree.fromstring(xml_bytes)
    paragraphs: list[str] = []
    for para in tree.iter(f"{{{_WORDML_NS}}}p"):
        texts = [
            t.text
            for t in para.iter(f"{{{_WORDML_NS}}}t")
            if t.text
        ]
        if texts:
            paragraphs.append("".join(texts))
    return (
        "[Contenido del archivo Word]\n\n" + "\n".join(paragraphs)
        if paragraphs
        else "[Archivo Word vacío]"
    )
