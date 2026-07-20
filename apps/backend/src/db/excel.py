"""Server-side .xlsx export for the SABBI portfolio.

Builds the workbook directly from Postgres data (via `ProductRepository`) —
no client-side spreadsheet library involved. One sheet per category with its
products, plus a "Portafolio Final" summary sheet consolidating all
categories, matching `portfolio-dashboard.spec.md` → "Exportar portafolio a
Excel".
"""

from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from db.models import Product

HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CURRENCY_FORMAT = '"$"#,##0'
PERCENT_FORMAT = "0.0%"

# Kept in sync with `apps/web/lib/categories.ts` (`CATEGORY_META`) and the
# `CATEGORIES` taxonomy in `agent/state.py`.
CATEGORY_ORDER: list[str] = [
    "inversiones_directas",
    "mercados_privados",
    "club_deals",
    "mercados_publicos",
    "otros",
    "cash_y_equivalentes",
]
CATEGORY_LABELS: dict[str, str] = {
    "inversiones_directas": "Inversiones directas",
    "mercados_privados": "Mercados privados",
    "club_deals": "Club deals",
    "mercados_publicos": "Mercados públicos",
    "otros": "Otros",
    "cash_y_equivalentes": "Cash y equivalentes",
}


def _style_header_row(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize_columns(ws, count: int, width: int = 26) -> None:
    for col_idx in range(1, count + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _composition_summary(product: Product) -> str:
    return ", ".join(f"{a.name} {a.percentage:.0f}%" for a in product.composition)


def _group_by_category(products: list[Product]) -> dict[str, list[Product]]:
    by_category: dict[str, list[Product]] = {}
    for product in products:
        by_category.setdefault(product.category, []).append(product)
    return by_category


def _write_category_sheet(wb: Workbook, category: str, products: list[Product]) -> None:
    label = CATEGORY_LABELS.get(category, category)
    ws = wb.create_sheet(title=label[:31])
    headers = ["Nombre", "Subcategoría", "Proveedor", "Monto (USD)", "Composición"]
    _style_header_row(ws, headers)

    row = 2
    for product in products:
        ws.cell(row=row, column=1, value=product.name)
        ws.cell(row=row, column=2, value=product.subcategory)
        ws.cell(row=row, column=3, value=product.provider)
        amount_cell = ws.cell(row=row, column=4, value=product.amount)
        amount_cell.number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=5, value=_composition_summary(product))
        row += 1

    total_label_cell = ws.cell(row=row, column=1, value="Total")
    total_label_cell.font = Font(bold=True)
    total_amount_cell = ws.cell(row=row, column=4, value=sum(p.amount for p in products))
    total_amount_cell.number_format = CURRENCY_FORMAT
    total_amount_cell.font = Font(bold=True)

    _autosize_columns(ws, len(headers))


def _write_summary_sheet(wb: Workbook, products: list[Product]) -> None:
    ws = wb.create_sheet(title="Portafolio Final", index=0)
    headers = ["Categoría", "Monto (USD)", "% del portafolio", "# Productos"]
    _style_header_row(ws, headers)

    total = sum(p.amount for p in products)
    by_category = _group_by_category(products)

    row = 2
    for category in CATEGORY_ORDER:
        category_products = by_category.get(category, [])
        if not category_products:
            continue
        category_total = sum(p.amount for p in category_products)
        ws.cell(row=row, column=1, value=CATEGORY_LABELS.get(category, category))
        amount_cell = ws.cell(row=row, column=2, value=category_total)
        amount_cell.number_format = CURRENCY_FORMAT
        percent_cell = ws.cell(
            row=row, column=3, value=(category_total / total if total else 0)
        )
        percent_cell.number_format = PERCENT_FORMAT
        ws.cell(row=row, column=4, value=len(category_products))
        row += 1

    total_label_cell = ws.cell(row=row, column=1, value="Total")
    total_label_cell.font = Font(bold=True)
    total_amount_cell = ws.cell(row=row, column=2, value=total)
    total_amount_cell.number_format = CURRENCY_FORMAT
    total_amount_cell.font = Font(bold=True)
    total_percent_cell = ws.cell(row=row, column=3, value=1.0 if total else 0)
    total_percent_cell.number_format = PERCENT_FORMAT
    total_percent_cell.font = Font(bold=True)
    total_count_cell = ws.cell(row=row, column=4, value=len(products))
    total_count_cell.font = Font(bold=True)

    _autosize_columns(ws, len(headers))


def build_portfolio_workbook(products: list[Product]) -> io.BytesIO:
    """Build the SABBI export workbook: a "Portafolio Final" summary sheet
    (index 0) followed by one sheet per category that has at least one
    product, and return it as an in-memory buffer ready for streaming.
    """
    wb = Workbook()
    wb.remove(wb.active)  # drop openpyxl's default blank sheet

    _write_summary_sheet(wb, products)

    by_category = _group_by_category(products)
    for category in CATEGORY_ORDER:
        category_products = by_category.get(category)
        if category_products:
            _write_category_sheet(wb, category, category_products)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_filename() -> str:
    return f"portafolio-sabbi-{date.today().isoformat()}.xlsx"
