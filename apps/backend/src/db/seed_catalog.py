"""Seed the product_catalog table from an Excel file.

Usage:
    python -m db.seed_catalog /path/to/products.xlsx
"""

from __future__ import annotations

import asyncio
import json
import re
import sys

from openpyxl import load_workbook

from db.connection import close_pool, get_pool

_ZW_RE = re.compile(r"[​‌‍﻿]")
_PCT_RE = re.compile(r"(.+?)\s+([\d.]+)%")


def _clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return _ZW_RE.sub("", text)


def _parse_underlying(text: str) -> str:
    """Parse 'Name1 X%, Name2 Y%' into JSON array of {name, percentage}."""
    cleaned = _clean(text)
    if not cleaned:
        return "[]"
    parts = [p.strip() for p in re.split(r",(?![^(]*\))", cleaned) if p.strip()]
    result = []
    for part in parts:
        part = part.strip().rstrip(",")
        m = _PCT_RE.match(part)
        if m:
            name = m.group(1).strip().rstrip(",")
            pct = float(m.group(2))
            result.append({"name": name, "percentage": pct})
    return json.dumps(result)


async def seed(path: str) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    pool = await get_pool()

    async with pool.acquire() as conn:
        await conn.execute("TRUNCATE product_catalog RESTART IDENTITY")

        count = 0
        for row in rows:
            if not row or not row[1]:
                continue
            name = _clean(row[1])
            if not name:
                continue
            await conn.execute(
                """INSERT INTO product_catalog
                   (name, geographic_focus, asset_class, underlying,
                    commission, currency, administrator, manager,
                    liquidity, return_rate)
                   VALUES ($1, $2, $3, $4::jsonb, $5, $6, $7, $8, $9, $10)""",
                name,
                _clean(row[2]),
                _clean(row[3]),
                _parse_underlying(row[4]),
                _clean(row[5]),
                _clean(row[6]),
                _clean(row[7]),
                _clean(row[8]),
                _clean(row[9]),
                _clean(row[10]) if len(row) > 10 else "",
            )
            count += 1

    await close_pool()
    return count


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python -m db.seed_catalog <path-to-excel>")
        sys.exit(1)
    total = asyncio.run(seed(sys.argv[1]))
    print(f"Seeded {total} products into product_catalog")
